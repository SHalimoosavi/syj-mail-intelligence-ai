"""
Regression test runner — calls process_email() directly for all 28 test
cases (bypassing real Gmail send/poll), then reads the actual DB results
and writes them into a copy of regression_test_suite.xlsx.

USAGE:
  1. Place regression_test_suite.xlsx in your project root (same folder
     as this script).
  2. pip install openpyxl   (inside your venv)
  3. python run_regression.py
  4. Open regression_test_suite_filled.xlsx — Actual Category, Actual
     Priority, and Draft Created are pre-filled. You still need to fill
     Reply Quality (1-5), Echo Detected, Pass/Fail, and Notes by reading
     the actual generated replies (printed to console as each test runs).

This does NOT touch real Gmail. It creates Email/Draft/Summary rows in
your actual database using fake gmail_id values (regression-test-01,
regression-test-02, ...) so they won't collide with real email IDs.
Safe to re-run; each run uses the same fake IDs so old test rows are
just re-processed (your Email table will accumulate rows across runs —
delete rows with gmail_id LIKE 'regression-test-%' between full runs if
you want a clean slate).
"""

import asyncio
from openpyxl import load_workbook

from app.workflows.pipeline import process_email
from app.core.database import get_session
from app.core.models import Email, Draft


class MockGmailClient:
    """No-op stand-in for the real Gmail client. Records what WOULD have
    happened without touching real Gmail."""

    def archive(self, gmail_id):
        print(f"    [mock] would archive {gmail_id}")

    def send_reply(self, thread_id, to, subject, body, in_reply_to_gmail_id):
        print(f"    [mock] would AUTO-SEND reply: {subject!r}")

    def create_draft(self, thread_id, to, subject, body):
        print(f"    [mock] would create Gmail draft: {subject!r}")


# Must match the Test Cases sheet row order exactly (row 2 = test_data[0]).
TEST_CASES = [
    ("ops@company.com", "URGENT: Production API is down",
     "The production API is unavailable. Please investigate immediately and let me know when service has been restored."),
    ("alerts@monitoring.company.com", "Database connection failures on prod-db-01",
     "We're seeing repeated connection timeouts on prod-db-01 since 2:14 AM. Customer-facing services are degraded. Need eyes on this ASAP."),
    ("support@company.com", "Client reporting their site is completely down",
     "Acme Corp just called — their storefront has been returning 500 errors for the last 20 minutes. They're losing sales. Can you check the deployment?"),
    ("security-noreply@github.com", "Security Alert: Unusual sign-in activity detected",
     "We detected a sign-in to your account from a new device in Lagos, Nigeria. If this was you, no action is needed. If not, please secure your account."),
    ("no-reply@service.com", "Your account has been temporarily locked",
     "For your security, we've locked your account after multiple failed login attempts. Reply to this email or visit your account settings to unlock it."),
    ("billing@vendorco.com", "Invoice #INV-20456 for June services",
     "Please find attached invoice #INV-20456 for services rendered in June, total $4,500, due within 30 days."),
    ("billing@vendorco.com", "Reminder: Invoice #INV-20456 is now overdue",
     "This is a friendly reminder that invoice #INV-20456 for $4,500 was due on July 1st and remains unpaid. Please arrange payment at your earliest convenience."),
    ("procurement@clientco.com", "New Purchase Order PO-8821",
     "Attached is purchase order PO-8821 for 50 units of Product X, requested delivery by August 15th."),
    ("legal@clientco.com", "Updated MSA — signature required",
     "Attached is the revised Master Service Agreement reflecting the new pricing terms we discussed. Please review and sign by end of week."),
    ("priya@clientco.com", "Confirming our call tomorrow at 3 PM",
     "Just confirming we're still on for our project review call tomorrow at 3 PM IST. I'll send the deck beforehand."),
    ("user123@customerco.com", "Bug: Export button crashes on Safari",
     "When I click 'Export to CSV' on Safari 17, the page freezes and I have to force-quit the browser. Works fine on Chrome. Can you take a look?"),
    ("user456@customerco.com", "Feature request: Dark mode",
     "Would love to see a dark mode option in the dashboard. Not urgent, just a nice-to-have for late-night work sessions."),
    ("angrycustomer@customerco.com", "Very disappointed with recent support experience",
     "I've emailed twice this week with no response and my issue still isn't resolved. This is affecting my team's ability to work. I need someone to address this today."),
    ("no-reply@service.com", "Password reset requested",
     "We received a request to reset the password for your account. Click the link below to choose a new password. This link expires in 1 hour."),
    ("notifications@github.com", "[repo] New comment on issue #142",
     "alice commented on issue #142: \"Can we prioritize this for the next sprint?\""),
    ("ci-noreply@github.com", "[repo] Build failed on main branch",
     "The build for commit a3f9c21 on branch 'main' failed during the test stage. 3 tests failed in auth_test.py. View logs for details."),
    ("notifications@github.com", "[repo] Pull Request #58: Fix memory leak in worker pool",
     "bob opened pull request #58 against main. 4 files changed, +120 -35. Requesting your review."),
    ("notifications@github.com", "[repo] Issue #201 assigned to you",
     "You've been assigned issue #201: \"Rate limiting not applied to /api/v2/search endpoint\"."),
    ("ads-noreply@linkedin.com", "Your campaign performance update is ready — act now to grow reach",
     "Your LinkedIn Ads campaign is performing well! Increase your budget today to keep the momentum going and reach even more prospects."),
    ("news@thehustle.co", "Today's top business stories",
     "A daily digest of curated business and tech news, handpicked by our editorial team."),
    ("events@saastool.com", "You're invited: Scaling AI Products — Free Webinar",
     "Join us next Thursday for a free webinar on scaling AI products from prototype to production. Register now to save your seat."),
    ("offers@saastool.com", "40% off your next renewal — today only",
     "As a valued customer, enjoy 40% off your next annual renewal. Offer expires tonight at midnight — don't miss out!"),
    ("personal-friend@gmail.com", "long time no talk!",
     "hey man, it's been forever! we should catch up soon, maybe grab chai this weekend if you're free?"),
    ("family-member@gmail.com", "Ammi's checkup went well",
     "Just wanted to let you know Ammi's checkup went well today, doctor said everything looks fine. Call when you get a chance."),
    ("friend@gmail.com", "You're invited to Zainab's birthday this Saturday!",
     "Come celebrate Zainab turning 30 this Saturday at 7 PM, our place. Let us know if you can make it!"),
    ("promo123@randomdeals.biz", "You've been selected for a FREE iPhone 17!!!",
     "Congratulations! You've been randomly selected to receive a FREE iPhone 17. Click here to claim your prize before it expires!"),
    ("security@paypa1-support.com", "Your account will be suspended — verify now",
     "We detected unusual activity on your account. To avoid suspension, please verify your identity immediately by clicking the link below and entering your login details."),
    ("unknown.official@mailservice.com", "Urgent business proposal — confidential",
     "I am a bank official with an urgent business proposal involving $12.5 million. I need your assistance to transfer these funds and will share 30% with you. Please reply with your bank details to proceed."),
]

assert len(TEST_CASES) == 28


async def run_all():
    mock_client = MockGmailClient()
    results = []

    for idx, (sender, subject, body) in enumerate(TEST_CASES):
        n = idx + 1
        gmail_id = f"regression-test-{n:02d}"

        print(f"\n[{n}/28] Processing: {subject}")

        parsed = {
            "gmail_id": gmail_id,
            "thread_id": f"thread-{gmail_id}",
            "sender": sender,
            "recipients": ["me@example.com"],
            "subject": subject,
            "body_text": body,
            "body_html": f"<p>{body}</p>",
            "has_attachments": False,
        }

        try:
            await process_email(mock_client, parsed)
        except Exception as exc:
            print(f"    !! process_email raised: {exc}")
            results.append({
                "n": n, "category": "ERROR", "priority": "ERROR",
                "draft": "N", "reply_body": "", "confidence": "",
                "reasoning": f"process_email crashed: {exc}",
            })
            continue

        with get_session() as session:
            email = session.query(Email).filter_by(gmail_id=gmail_id).first()

            if email is None:
                print("    !! No Email row found after processing.")
                continue

            draft = (
                session.query(Draft)
                .filter_by(email_id=email.id)
                .order_by(Draft.id.desc())
                .first()
            )

            result = {
                "n": n,
                "category": email.category,
                "priority": email.priority,
                "draft": "Y" if draft else "N",
                "reply_body": draft.reply_body if draft else "",
                "confidence": draft.confidence if draft else "",
                "reasoning": draft.reasoning if draft else "",
            }
            results.append(result)

            print(f"    category={result['category']} priority={result['priority']} draft={result['draft']}")
            if draft:
                print(f"    confidence={draft.confidence} status={draft.status}")
                print(f"    reply preview: {draft.reply_body[:150]}")

    return results


def write_to_spreadsheet(results, template="regression_test_suite.xlsx", output="regression_test_suite_filled.xlsx"):
    wb = load_workbook(template)
    ws = wb["Test Cases"]

    for r in results:
        row = r["n"] + 1  # row 2 = test 1

        ws.cell(row=row, column=9, value=r["category"])       # Actual Category
        ws.cell(row=row, column=10, value=r["priority"])       # Actual Priority
        ws.cell(row=row, column=11, value=r["draft"])          # Draft Created

        auto_note = ""
        if r["draft"] == "Y":
            auto_note = f"[AUTO] confidence={r['confidence']}; reasoning={r['reasoning']}; reply={r['reply_body'][:300]}"
        elif r["category"] == "ERROR":
            auto_note = f"[AUTO] {r['reasoning']}"

        ws.cell(row=row, column=17, value=auto_note)            # Notes

    wb.save(output)
    print(f"\nSaved results to {output}")
    print("Now open it and fill in: Reply Quality (1-5), Echo Detected, Pass/Fail")


if __name__ == "__main__":
    results = asyncio.run(run_all())
    write_to_spreadsheet(results)
